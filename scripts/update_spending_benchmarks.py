#!/usr/bin/env python3
import io
import json
import re
from pathlib import Path

import openpyxl
import requests

YEAR = "2023-2024"
INDEX_URL = "https://www.bls.gov/cex/tables/cross-tab/mean.htm"
REGIONS = {
    "northeast": "Northeast",
    "midwest": "Midwest",
    "south": "South",
    "west": "West",
}
BANDS = [
    {"id": "lt15", "label": "Less than $15,000", "min": 0, "max": 14999},
    {"id": "15-30", "label": "$15,000 to $29,999", "min": 15000, "max": 29999},
    {"id": "30-40", "label": "$30,000 to $39,999", "min": 30000, "max": 39999},
    {"id": "40-50", "label": "$40,000 to $49,999", "min": 40000, "max": 49999},
    {"id": "50-70", "label": "$50,000 to $69,999", "min": 50000, "max": 69999},
    {"id": "70-100", "label": "$70,000 to $99,999", "min": 70000, "max": 99999},
    {"id": "100-150", "label": "$100,000 to $149,999", "min": 100000, "max": 149999},
    {"id": "150-200", "label": "$150,000 to $199,999", "min": 150000, "max": 199999},
    {"id": "200plus", "label": "$200,000 and more", "min": 200000, "max": None},
]
FIELDS = {
    "average annual expenditures": "annualExpenditures",
    "income before taxes": "averageIncomeBeforeTax",
    "income after taxes": "averageIncomeAfterTax",
    "people": "averagePeople",
    "food": "food",
    "housing": "housing",
    "apparel and services": "apparel",
    "transportation": "transportation",
    "healthcare": "healthcare",
    "entertainment": "entertainment",
    "personal care products and services": "personalCare",
    "education": "education",
    "miscellaneous": "miscellaneous",
    "personal insurance and pensions": "personalInsurancePensions",
}

BROWSER_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/151.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
}


def text(value):
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def numeric_tail(row, count=10):
    values = []
    for value in row:
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            values.append(float(value))
    if len(values) < count:
        return None
    return values[-count:]


def row_label(row):
    candidates = [text(value) for value in row[:8] if isinstance(value, str) and value.strip()]
    return " ".join(candidates)


def find_field(rows, needle):
    matches = []
    for row in rows:
        label = row_label(row)
        if needle in label:
            nums = numeric_tail(row)
            if nums:
                matches.append((label, nums))
    if not matches:
        raise RuntimeError(f"Could not find row for {needle!r}")
    matches.sort(key=lambda item: len(item[0]))
    return matches[0][1]


def open_bls_session():
    session = requests.Session()
    session.headers.update(BROWSER_HEADERS)
    warm = session.get(INDEX_URL, timeout=60)
    warm.raise_for_status()
    return session


def extract_region(session, slug, region_name):
    url = f"https://www.bls.gov/cex/tables/cross-tab/mean/cu-region-by-income-{slug}-{YEAR}.xlsx"
    headers = {
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream;q=0.9,*/*;q=0.8",
        "Referer": INDEX_URL,
    }
    response = session.get(url, timeout=60, headers=headers)
    response.raise_for_status()
    if not response.content.startswith(b"PK"):
        raise RuntimeError(f"BLS did not return an Excel workbook for {region_name}")
    workbook = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))

    extracted = {field: find_field(rows, needle) for needle, field in FIELDS.items()}
    cohorts = []
    for index, band in enumerate(BANDS, start=1):
        cohort = dict(band)
        for field, values in extracted.items():
            cohort[field] = values[index]
        cohorts.append(cohort)

    return {
        "name": region_name,
        "source": url,
        "cohorts": cohorts,
    }


def main():
    session = open_bls_session()
    regions = {slug: extract_region(session, slug, name) for slug, name in REGIONS.items()}
    result = {
        "schemaVersion": 1,
        "source": "U.S. Bureau of Labor Statistics Consumer Expenditure Surveys",
        "sourcePeriod": YEAR,
        "generatedFrom": "BLS region of residence by income before taxes cross-tabulated means tables",
        "generatedAt": __import__("datetime").datetime.now(__import__("datetime").timezone.utc).isoformat(),
        "regions": regions,
        "notes": [
            "Peer cohorts are descriptive population means, not recommended budgets.",
            "BLS suppresses unreliable published estimates; ShareCapsule should not treat a peer mean as a universal target.",
            "The health engine compares peer patterns with the user's own cash flow, reserves and debt before suggesting a direction.",
        ],
    }
    out = Path("health/benchmarks.json")
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(result, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "sourcePeriod": YEAR,
        "regions": {key: len(value["cohorts"]) for key, value in regions.items()},
        "output": str(out),
    }))


if __name__ == "__main__":
    main()
